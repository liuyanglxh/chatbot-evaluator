"""
Excel批量评估处理核心逻辑
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from font_utils import font_manager


class ExcelEvaluationHandler:
    """Excel评估处理器"""

    def __init__(self, parent_window, excel_file_path, evaluators, config_manager):
        self.parent_window = parent_window
        self.excel_file_path = excel_file_path
        self.evaluators = evaluators
        self.config_manager = config_manager

        # 创建进度窗口
        self.progress_window = None
        self.progress_var = None
        self.status_label = None

    def run(self):
        """运行批量评估"""
        # 创建进度窗口
        self.create_progress_window()

        # 在后台线程执行
        thread = threading.Thread(target=self._execute_evaluation)
        thread.daemon = True
        thread.start()

    def create_progress_window(self):
        """创建进度窗口"""
        self.progress_window = tk.Toplevel(self.parent_window)
        self.progress_window.title("正在评估")
        self.progress_window.geometry("500x200")
        self.progress_window.transient(self.parent_window)
        self.progress_window.grab_set()
        self.progress_window.resizable(False, False)

        # 居中显示
        self.progress_window.update_idletasks()
        width = self.progress_window.winfo_width()
        height = self.progress_window.winfo_height()
        screen_width = self.progress_window.winfo_screenwidth()
        screen_height = self.progress_window.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.progress_window.geometry(f'{width}x{height}+{x}+{y}')

        # 内容
        frame = ttk.Frame(self.progress_window, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        ttk.Label(
            frame,
            text="⏳ 正在批量评估，请稍候...",
            font=font_manager.panel_title_font()
        ).pack(pady=(0, 20))

        # 进度条
        self.progress_var = tk.DoubleVar(value=0)
        progress_bar = ttk.Progressbar(
            frame,
            variable=self.progress_var,
            maximum=100,
            length=400
        )
        progress_bar.pack(pady=(0, 10))

        # 状态标签
        self.status_label = ttk.Label(
            frame,
            text="准备中...",
            font=font_manager.panel_font()
        )
        self.status_label.pack()

    def _execute_evaluation(self):
        """执行评估（后台线程）"""
        try:
            # 1. 解析Excel
            self._update_status("正在解析Excel文件...", 10)
            conversations = self._parse_excel()

            if not conversations:
                self._show_error("Excel文件中没有找到有效数据")
                return

            # 2. 批量评估
            total_conversations = len(conversations)
            total_evaluators = len(self.evaluators)
            total_tasks = total_conversations * total_evaluators

            current_task = 0

            # 存储所有评估结果
            all_results = {}

            for conv_id, turns in conversations.items():
                all_results[conv_id] = {
                    'turns': turns,
                    'results': {}
                }

                for evaluator in self.evaluators:
                    current_task += 1
                    progress = 10 + (current_task / total_tasks) * 80
                    self._update_status(
                        f"正在评估对话 {conv_id} ({len(turns)}轮) - {evaluator['name']}...",
                        progress
                    )

                    # 评估
                    evaluator_results = self._evaluate_conversation(
                        turns,
                        evaluator
                    )
                    all_results[conv_id]['results'][evaluator['name']] = evaluator_results

            # 3. 写入Excel结果
            self._update_status("正在生成Excel评估结果...", 90)
            excel_output_path = self._write_results_to_excel(all_results)

            # 4. 生成HTML报告
            self._update_status("正在生成HTML报告...", 95)
            html_output_path = self._generate_html_report(all_results)

            # 5. 完成
            self._update_status("评估完成！", 100)

            # 显示成功消息（传递两个文件路径）
            self.progress_window.after(0, self._show_success, excel_output_path, html_output_path)

        except Exception as e:
            import traceback
            error_msg = f"评估失败: {str(e)}\n\n{traceback.format_exc()}"
            self.progress_window.after(0, self._show_error, error_msg)

    def _parse_excel(self):
        """解析Excel文件，按编号分组对话（支持期望回答）"""
        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb[wb.sheetnames[0]]  # 读取第一个sheet

        conversations = {}
        current_id = None

        # 从第2行开始读取（第1行是表头）
        for row_idx in range(2, ws.max_row + 1):
            col_id = ws.cell(row_idx, 1).value  # 编号列
            question = ws.cell(row_idx, 2).value or ""  # 问题列
            answer = ws.cell(row_idx, 3).value or ""  # 回答列
            context = ws.cell(row_idx, 4).value or ""  # 参考资料列
            expected_answer = ws.cell(row_idx, 5).value or ""  # 期望回答列（可选，第5列）

            # 如果有编号，开始新对话
            if col_id:
                current_id = str(col_id)
                conversations[current_id] = []

            # 添加到当前对话
            if current_id:
                conversations[current_id].append({
                    'question': str(question).strip(),
                    'answer': str(answer).strip(),
                    'context': str(context).strip(),
                    'expected_answer': str(expected_answer).strip(),  # 添加期望回答字段
                    'row_index': row_idx
                })

        return conversations

    def _evaluate_conversation(self, turns, evaluator):
        """评估单个对话（可能是单轮或多轮）"""
        from evaluators import get_executor

        model_settings = self.config_manager.get_model_settings()
        executor = get_executor(evaluator)

        turn_mode = evaluator.get('turn_mode', 'single')
        results = []

        if turn_mode == 'single':
            # 单轮评估器：每轮独立评估
            for i, turn in enumerate(turns):
                # 构建累积上下文
                context_parts = []

                # 历史轮次
                for j in range(i):
                    hist_turn = turns[j]
                    context_parts.append(f"第{j+1}轮:")
                    context_parts.append(f"问题：{hist_turn['question']}")
                    context_parts.append(f"回答：{hist_turn['answer']}")
                    if hist_turn['context']:
                        context_parts.append(f"参考资料：{hist_turn['context']}")
                    context_parts.append("")

                # 当前轮次的参考资料
                if turn['context']:
                    context_parts.append(f"第{i+1}轮:")
                    context_parts.append(f"参考资料：{turn['context']}")

                full_context = "\n".join(context_parts)

                # 获取期望回答（可选）
                expected_answer = turn.get('expected_answer', '').strip()

                # 执行评估（传入期望回答）
                result = executor.execute(
                    turn['question'],
                    turn['answer'],
                    full_context,
                    model_settings,
                    expected_answer  # 传入期望回答
                )

                results.append({
                    'turn_index': i,
                    'score': result.get('score', 0),
                    'reason': result.get('reason', ''),
                    'success': result.get('success', False),
                    'merged': False  # 单轮评估器不合并
                })
        else:
            # 多轮评估器：整体评估一次
            # 构建完整对话文本
            conversation_parts = []
            for i, turn in enumerate(turns, 1):
                question = turn['question'].strip()
                answer = turn['answer'].strip()
                context = turn['context'].strip()

                turn_text = f"第{i}轮:\n问题: {question}\n回答: {answer}"
                if context:
                    turn_text += f"\n参考资料: {context}"
                turn_text += "\n"

                conversation_parts.append(turn_text)

            full_conversation = "\n".join(conversation_parts)

            # 对于多轮评估，如果有任意一轮有期望回答，则使用第一个非空的期望回答
            expected_answer = None
            for turn in turns:
                ea = turn.get('expected_answer', '').strip()
                if ea:
                    expected_answer = ea
                    break

            # 执行评估（传入期望回答）
            result = executor.execute(
                full_conversation,
                "",  # 多轮模式下answer为空
                "",
                model_settings,
                expected_answer  # 传入期望回答
            )

            # 所有轮次共享同一个结果
            for i in range(len(turns)):
                results.append({
                    'turn_index': i,
                    'score': result.get('score', 0),
                    'reason': result.get('reason', ''),
                    'success': result.get('success', False),
                    'merged': True  # 多轮评估器需要合并
                })

        return results

    def _generate_html_report(self, all_results):
        """生成HTML报告"""
        from html_report_generator import HtmlReportGenerator

        # 生成HTML文件路径
        excel_path = Path(self.excel_file_path)
        html_filename = f"{excel_path.stem}_评估报告.html"
        html_output_path = excel_path.parent / html_filename

        # 创建报告生成器
        generator = HtmlReportGenerator()

        # 生成报告
        generator.generate_report(
            all_results=all_results,
            evaluators=self.evaluators,
            output_path=str(html_output_path)
        )

        return str(html_output_path)

    def _write_results_to_excel(self, all_results):
        """将结果写入Excel"""
        wb = openpyxl.load_workbook(self.excel_file_path)

        # 创建新的sheet
        if "评估结果" in wb.sheetnames:
            wb.remove(wb["评估结果"])

        ws_result = wb.create_sheet("评估结果")

        # 复制原Sheet1的数据到新sheet
        ws_original = wb[wb.sheetnames[0]]

        # 复制表头（包括第5列期望回答）
        for col_idx in range(1, 6):  # 1-5列：编号、问题、回答、参考资料、期望回答
            cell = ws_original.cell(1, col_idx)
            ws_result.cell(1, col_idx, cell.value)

        # 复制数据行
        row_mapping = {}  # 记录每个对话在新sheet中的行范围
        current_row = 2

        for conv_id, data in all_results.items():
            turns = data['turns']
            start_row = current_row

            for turn in turns:
                original_row = turn['row_index']
                for col_idx in range(1, 6):  # 复制5列数据，包括期望回答
                    cell = ws_original.cell(original_row, col_idx)
                    ws_result.cell(current_row, col_idx, cell.value)
                current_row += 1

            end_row = current_row - 1
            row_mapping[conv_id] = (start_row, end_row)

        # 添加评估器列（从第6列开始，前5列是原数据）
        current_col = 6
        for evaluator in self.evaluators:
            evaluator_name = evaluator['name']

            # 添加分数列
            score_col_letter = get_column_letter(current_col)
            ws_result.cell(1, current_col, evaluator_name)

            # 添加原因列
            reason_col_letter = get_column_letter(current_col + 1)
            ws_result.cell(1, current_col + 1, "原因")

            # 填充数据
            for conv_id, data in all_results.items():
                start_row, end_row = row_mapping[conv_id]
                results = data['results'][evaluator_name]

                if results and results[0]['merged']:
                    # 多轮评估器：合并单元格
                    ws_result.merge_cells(f"{score_col_letter}{start_row}:{score_col_letter}{end_row}")
                    ws_result.merge_cells(f"{reason_col_letter}{start_row}:{reason_col_letter}{end_row}")

                    # 只在第一行填值
                    ws_result.cell(start_row, current_col, results[0]['score'])
                    ws_result.cell(start_row, current_col + 1, results[0]['reason'])
                else:
                    # 单轮评估器：每行单独填值
                    for i, result in enumerate(results):
                        row = start_row + i
                        ws_result.cell(row, current_col, result['score'])
                        ws_result.cell(row, current_col + 1, result['reason'])

            current_col += 2

        # 保存文件
        output_path = Path(self.excel_file_path).parent / f"{Path(self.excel_file_path).stem}_评估结果.xlsx"
        wb.save(str(output_path))

        return str(output_path)

    def _update_status(self, status, progress):
        """更新进度窗口状态"""
        if self.progress_window and self.progress_window.winfo_exists():
            self.progress_window.after(0, lambda: self._do_update_status(status, progress))

    def _do_update_status(self, status, progress):
        """实际更新状态的函数"""
        if self.status_label:
            self.status_label.config(text=status)
        if self.progress_var:
            self.progress_var.set(progress)

    def _show_success(self, excel_output_path, html_output_path):
        """显示成功消息"""
        if self.progress_window and self.progress_window.winfo_exists():
            self.progress_window.destroy()

        result = messagebox.askyesno(
            "评估完成",
            f"✅ 评估完成！\n\n结果已保存:\n"
            f"📊 Excel: {Path(excel_output_path).name}\n"
            f"📄 HTML报告: {Path(html_output_path).name}\n\n"
            f"是否在浏览器中打开HTML报告？"
        )

        if result:
            self._open_html_report(html_output_path)
        else:
            # 如果用户选择不打开，询问是否打开文件夹
            result2 = messagebox.askyesno(
                "打开文件夹",
                "是否打开文件所在文件夹？"
            )
            if result2:
                import subprocess
                import platform
                folder_path = str(Path(html_output_path).parent)

                if platform.system() == 'Windows':
                    subprocess.Popen(f'explorer /select,"{html_output_path}"')
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.Popen(['open', '-R', html_output_path])
                else:  # Linux
                    subprocess.Popen(['xdg-open', folder_path])

    def _open_html_report(self, html_path):
        """在浏览器中打开HTML报告"""
        import subprocess
        import platform
        import webbrowser

        try:
            # 尝试使用系统默认浏览器
            if platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', html_path])
            elif platform.system() == 'Windows':
                os.startfile(html_path)  # type: ignore
            else:  # Linux
                subprocess.Popen(['xdg-open', html_path])
        except Exception as e:
            # 降级到 webbrowser
            print(f"使用系统命令打开失败，尝试使用 webbrowser: {e}")
            webbrowser.open('file://' + html_path)

    def _show_error(self, error_msg):
        """显示错误消息"""
        if self.progress_window and self.progress_window.winfo_exists():
            self.progress_window.destroy()

        messagebox.showerror("错误", error_msg)
